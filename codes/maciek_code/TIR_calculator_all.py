# -*- coding: utf-8 -*-
"""
Created on Thu Oct  8 13:45:20 2020

@author: hol428
"""

import pandas as pd
import glob
import re
from openpyxl import load_workbook

def create_gfpod():
    GFPOD = OD.copy(deep=True)
    i=0
    
    OD.drop('Time',axis=1,inplace=True)
    
    for col in OD.columns:
        for ind in GFPOD.index:
            # print(col)
            # print(ind)
            GFPOD[col][ind]=GFP[col][ind]/OD[col][ind]
            
    for ind in GFPOD.index:
        GFPOD['Time'][ind]=i
        i+=10
    
    GFPOD.set_index('Time',inplace=True)
    
    with pd.ExcelWriter(File, engine="openpyxl", mode='a') as writer:  
        GFPOD.to_excel(writer, sheet_name='GFPOD')

def create_tir():
    GFPOD.drop('Time',axis=1,inplace=True)
    # GFPOD.reset_index(drop=True,inplace=True)
        
    minp = GFPOD.min(axis=0)
    
    time = [240]
    
    rates = pd.DataFrame(columns=['240','Comment'],index = list(GFPOD.columns))
    
    for t in time:
        for column,value in minp.iteritems():
            start = GFPOD[GFPOD[column]==value].index.values.astype(int)[0]
            end = start + (t/10)
            rate = (GFPOD.loc[start:end, [column]].sum() / t)
            # print(start)
            # print(end)
            # print(rate)
            rates.at[column,str(t)] = float(rate[0])
            if rate[0]<0:
                rates.at[column,'Comment'] = 'Impossible value'
            elif rate[0]>100 and OD.max(axis=0)[column]<0.1:          
                rates.at[column,'Comment'] = 'Possibly an empty well'
            else:
                rates.at[column,'Comment'] = 'Correct value'
    
    print(rates)
    
    with pd.ExcelWriter(File, engine="openpyxl", mode='a') as writer:  
        rates.to_excel(writer, sheet_name='TIR')

path = '../../data/Plate_results/*.xlsx'
ResFile = '../../data/Results_Masterfile_test.xlsx'
Results = pd.read_excel(ResFile,sheet_name='Microplate',encoding="utf-8-sig") 
print(Results.head())
for File in glob.glob(path):
    
    print(File)
    
    if "Rep" in File :
        OD = pd.read_excel(File,sheet_name='OD',encoding="utf-8-sig")
        GFP = pd.read_excel(File,sheet_name='GFP',encoding="utf-8-sig")
        wb = load_workbook(File, read_only=True)   # open an Excel file and return a workbook
    
        if 'GFPOD' not in wb.sheetnames:
            create_gfpod()
    
        if 'TIR' not in wb.sheetnames:
            GFPOD = pd.read_excel(File,sheet_name='GFPOD')
            create_tir()
    
    # TIR = pd.read_excel(File,sheet_name='TIR')
    # rep = re.search(r'Rep\d',File).group(0)
    # plate = re.search(r'\\.+Plate',File).group(0)[1:]
    # print(rep + ' ' + plate)
    
    
    # z = Results[rep].where(Results['Plate'] == plate).first_valid_index()   
    # print (z)
    # print(Results[rep].where(Results['Plate'] == plate))
    # # print(Results[rep][z])
    # # if Results[rep].where(Results['Plate'] == plate).dropna().empty:
    # # if z is not None:
    #     # for i,j in enumerate(Results[rep].where(Results['Plate'] == plate)):
    #     #     if i>90:
    #     #         break
    #     #     Results.replace((Results[rep][z]),TIR['240'][i],inplace=True)
    #     #     z=+1
        
    # if z is None:
    #     if plate == "First_Plate":            
    #         for i,j in enumerate(Results[rep].where(Results['Plate'] == plate)):
    #             if i>90:
    #                 break
    #             Results[rep][i] = TIR['240'][i]
                
    #     if plate == "Second_Plate":  
                            
    #         for i,j in enumerate(Results[rep].where(Results['Plate'] == plate)):
    #             if i>90:
    #                 break
    #             i += 90
    #             Results[rep][i] = TIR['240'][i]
                
    #     if plate == "Third_Plate":            
    #         for i,j in enumerate(Results[rep].where(Results['Plate'] == plate)):
    #             if i>90:
    #                 break
    #             d = i + 178
    #             Results[rep][d] = TIR['240'][i]
            
# with pd.ExcelWriter(ResFile, engine="openpyxl", mode='a') as writer:  
#     Results.to_excel(writer, sheet_name='test')
